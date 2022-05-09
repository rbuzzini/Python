import openpyxl

# import your data
wb = openpyxl.load_workbook(r'C:\Users\rbuzzini\Documents\Personale\Git\Python\transaction.xlsx')

sh1 = wb['Sheet1']  # to save sheetName we want

# sh1 is a object of openpyxl.woksheet.worksheet.Worksheet class:
type(sh1)

# to access specific cell values from the worksheet:
sh1['D2'].value   # option 1
sh1.cell(2, 4).value   # option 2


# get_sheet_by_name method:
wb.get_sheet_by_name('Sheet1')['D2'].value

# to access the worksheet rows:
sh1.max_row

# to access worksheet columns:
sh1.max_column

# for loop to print all values from a specific worksheet:

for i in range(1, sh1.max_row+1):
    for j in (1, sh1.max_column+1):
        print(sh1.cell(i, j).value)



from openpyxl import Workbook

wb2 = Workbook()   # object of workbook class. It is like creating an excel file

wb2.active   # it returns the active sheet

wb2.sheetnames   # sheetnames in the excel file

wb2['Sheet'].title = 'Automation Report'   # to change sheetname
sh2 = wb2.active

# Let's enter some data:
sh2['A1'] = 'Name'
sh2['B1'] = 'Gender'
sh2['A2'] = 'John'
sh2['B2'] = 'Male'
sh2['A3'] = 'Jennifer'
sh2['B3'] = 'Female'

wb2.save(r'FinalReport.xlsx')   # to save the excel file