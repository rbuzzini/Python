# Let's import all the libraries we'll need along the tutorial:
from openpyxl import Workbook
from openpyxl import load_workbook
import json
from datetime import datetime

from classes import Transaction
from mapping import TXN_ID, MEMBER_ID, TICKER, TXN_DATE, TXN_TIME, TXN_TYPE, QUANTITY, PERCENTAGE_FEE


# region 1 - Reading Excel Spreadsheets With openpyxl
#  Let’s start reading some spreadsheets! To begin with, open our sample spreadsheet:
wb = load_workbook(r'C:\Users\rbuzzini\Documents\Personale\Git\Python\transaction.xlsx')

# Workbook sheetnames:
wb.sheetnames

# Get active sheet
sheet = wb.active

# See sheet title:
sheet.title


# Now, after opening a spreadsheet, you can easily retrieve data from it like this:
sheet['A1']
sheet['A1'].value



"""
To return the actual value of a cell, you need to do .value. Otherwise, 
you’ll get the main Cell object. You can also use the method .cell() to retrieve
a cell using index notation. Remember to add .value to get the actual value 
and not a Cell object:
"""
sheet.cell(row=10, column=5)

# 1.2 - Additional reading options
"""
There are a few arguments you can pass to load_workbook() that change the way 
a spreadsheet is loaded. The most important ones are the following two Booleans:
1- `read_only` loads a spreadsheet in read-only mode allowing you to open very large Excel files.
2 - `data_only` ignores loading formulas and instead loads only the resulting values.
"""


# 1.3 - Importing Data From a Spreadsheet

# Iterating through the Data
"""
There are a few different ways you can iterate through the data depending on your needs.

You can slice the data with a combination of columns and rows:
"""

sheet['A1:H5']

# Get all cells from column A
sheet['A']

# Get all cells for a range of columns:
sheet['A:B']

# Get all the cells from row 10:
sheet[10]

# Get all cells for a range of rows:
sheet[10:15]

# You’ll notice that all of the above examples return a `tuple`.

"""
There are also multiple ways of using normal Python generators to go through the data. The main methods you can use to achieve this are:
    1 - .iter_rows()
    2 - .iter_cols()

Both methods can receive the following arguments:
    - min_row
    - max_row
    - min_col
    - max_col
"""

# These arguments are used to set boundaries for the iteration:

for row in sheet.iter_rows(min_row=1,
                           max_row=5, 
                           min_col=1, 
                           max_col=8):
    print(row)

for column in sheet.iter_cols(min_row=1,
                           max_row=5, 
                           min_col=1, 
                           max_col=8):
    print(column)


"""
You’ll notice that in the first example, when iterating through the rows using 
.iter_rows(), you get one tuple element per row selected. While when using 
.iter_cols() and iterating through columns, you’ll get one tuple per column instead.

One additional argument you can pass to both methods is the Boolean values_only. 
When it’s set to True, the values of the cell are returned, instead of the Cell object:
"""

for row in sheet.iter_rows(min_row=1,
                           max_row=5, 
                           min_col=1, 
                           max_col=8,
                           values_only=True):
    print(row)


# Nice! Now that you know how to get all the important product information 
# you need, let’s put that data into a dictionary:

transactions = {}


for row in sheet.iter_rows(min_row=2,
                           max_row=sheet.max_row,
                           # min_col=1,
                           max_col=sheet.max_column,
                           values_only=True):
    
    txn_id = row[0]
    transaction = {
        'member_id': row[1],
        'ticker': row[2],
        'txn_date': row[3],
        'txn_time': row[4],
        'txn_type': row[5],
        'quantity': row[6],
        'percentage_fee': row[7]
    }
    transactions[txn_id] = transaction

# Using json here to be able to format the output for displaying later,
# json.dumps() function converts a Python object into a json string
print(json.dumps(transactions))

# endregion


# region 2 - Convert Data Into Python Classes

"""
To finalize the reading section of this tutorial, let’s dive into Python
classes and see how you could improve on the example above and better structure the data.
So, first things first, let’s look at the data you have and decide what you 
want to store and how you want to store it.
"""

# see classes.py file


# Let’s create a file mapping.py where you have a list of all the field names 
# and their column location (zero-indexed) on the spreadsheet:
# see mapping.py file

# You don’t necessarily have to do the mapping above. It’s more for readability 
# when parsing the row data, so you don’t end up with a lot of magic numbers lying around.

# Finally, let’s look at the code needed to parse the spreadsheet data into a list of transactions:

# Using the values_only because you just want to return the cell value

transactions = []
for row in sheet.iter_rows(min_row=2, values_only=True):
     
    # converting string to dates
    txnDate = datetime.strptime(row[TXN_DATE], '%Y-%m-%d')
    txnTime = datetime.strptime(row[TXN_TIME], '%Y-%m-%d %H:%M:%S')

    transaction = Transaction(
        txn_id=row[TXN_ID],
        member_id=row[MEMBER_ID],
        ticker=row[TICKER],
        txn_date=txnDate,
        txn_time=txnTime,
        txn_type=row[TXN_TYPE],
        quantity=row[QUANTITY],
        percentage_fee=row[PERCENTAGE_FEE])
    transactions.append(transaction)

transactions[0]

"""
That’s it! Now you should have the data in a very simple and digestible class format,
and you can start thinking of storing this in a Database or any other type of data storage you like.

Using this kind of OOP strategy to parse spreadsheets makes handling the data much simpler later on.
"""

# endregion


# region 3 - Writing Excel Spreadsheets With openpyxl

# Appending New Data

# Start by opening the spreadsheet and selecting the main sheet

# Write what you want into a specific cell
# sheet["C1"] = "writing ;)"

# Save the spreadsheet
# workbook.save(filename="hello_world_append.xlsx")


# There are a lot of different things you can write to a spreadsheet, from 
# simple text or number values to complex formulas, charts, or even images.

# Let’s start creating some spreadsheets!

# Creating a Simple Spreadsheet:
filename = 'hello_world.xlsx'
workbook = Workbook()
sheet = workbook.active

sheet['A1'] = 'Hello'
sheet['B1'] = 'World!'

workbook.save(filename=filename)


# ...


# endregion


# See more on:
# https://realpython.com/openpyxl-excel-spreadsheets-python/